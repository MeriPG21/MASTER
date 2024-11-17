import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def estadistica_juego():
    wb = None
    try:
        wb = openpyxl.load_workbook("estadisticas.xlsx", data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=3, values_only=True):
            jugador, resultado, juego = row
            print(f"Jugador: {jugador}, Resultado: {resultado}, Juego: {juego}")
        wb.save("estadisticas.xlsx")
    except FileNotFoundError:
        print("Aún no hay estadísticas disponibles.")
    except (InvalidFileException) as e:
        print(f"Error al leer estadísticas: {e}")
    finally:
        if wb:
            wb.close()

def registrar_resultado(jugador, resultado, juego):
    wb = None
    try:
        wb = openpyxl.load_workbook("estadisticas.xlsx", data_only=True)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.append(["Jugador", "Resultado", "Juego"])

    sheet = wb.active
    sheet.append([jugador, resultado, juego])
    
    try:
        wb.save("estadisticas.xlsx")
    except (InvalidFileException) as e:
        print(f"Error al guardar estadísticas: {e}")
    finally:
        if wb:
            wb.close()
